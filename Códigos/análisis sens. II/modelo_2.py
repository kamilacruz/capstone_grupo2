# modelo.py
from gurobipy import GRB, Model, quicksum, max_
from variables_parametros2 import ModeloVariablesParametros
import os
import pandas as pd 
import parametros as prm
from openpyxl import load_workbook

def correr_modelo():
    modelo = ModeloVariablesParametros()
    m = modelo.obtener_modelo()

    #variables ----- ojo con el orden
    q, p, z, v, inv, di, DE = modelo.obtener_variables()

    #parametros 
    C_PROD, C_FIJO, C_INV, CAPACIDAD_T1, CAPACIDAD_T2, CMO, d, A, B1, B2, d_estimada, proporcion_maxima = modelo.obtener_parametros()
    gamma = modelo.gamma

    M = 1000000
    rangos = modelo.obtener_rangos()
    I_ = rangos['I_']
    J_ = rangos['J_']
    T_ = rangos['T_']

    #funcion objetivo 
    
    FO = quicksum(p[i, j, t] * v[i, j, t] - C_INV[i] * inv[i, j, t] - C_PROD[i] * q[i, j, t] - C_FIJO[i] * z[i, j, t] - 0.1 * di[i, j, t] * p[i, j, t] for i in I_ for j in J_ for t in T_)
    

    m.setObjective(FO, GRB.MAXIMIZE) 

    # 1. Qijt <= M*zijt	
    m.addConstrs(q[i, j, t] <= M * z[i, j, t] for i in I_ for j in J_ for t in T_)

    # 2. Pijt >= 1.05(CPi)
    m.addConstrs(p[i, j, t] >= 1.05 * C_PROD[i] for i in I_ for j in J_ for t in T_)

    # 3. No arbitraje Pi1t – Pi2t <= 3.8, Pi2t – Pi1t <= 3.8
    m.addConstrs(p[i, 1, t] - p[i, 2, t] <= 3800 for i in I_ for t in T_)
    m.addConstrs(p[i, 2, t] - p[i, 1, t] <= 3800 for i in I_ for t in T_)

    # 4. Sum{i=1}_{N} Ii1t <= 175000 Sum{i=1}_{N} Ii2t <= 163000
    m.addConstrs(quicksum(inv[i, 1, t] for i in I_) <= CAPACIDAD_T1 for t in T_)
    m.addConstrs(quicksum(inv[i, 2, t] for i in I_) <= CAPACIDAD_T2 for t in T_)

    # 5. DE_ijt = (d_estimada)_ijt + gamma_ij * (P_ijt - P_ijt-1)
    # Verificar qe d_estimada es la demanda que nos da el ML
    # Falta definir el gamma. Esto de debería hacer en el archivo variables_parametros2. Tengo entendido que sería parámetro(?
   # m.addConstrs(
    #    DE[i, j, t] == (
     #       d_estimada[i, j, t] if t == prm.T_INICIO
      #      else d_estimada[i, j, t] + gamma[i, j] * (p[i, j, t] - p[i, j, t - 1])
       # )
        #for i in I_ for j in J_ for t in T_ if t >= prm.T_INICIO  # <- esto está bien
    #)
    # Para t = T_INICIO
    m.addConstrs(
        (DE[i, j, prm.T_INICIO] == d_estimada[f"{i},{j},{prm.T_INICIO}"]
        for i in I_ for j in J_), name="def_DE_inicio"
    )

    # Para t > T_INICIO
    m.addConstrs(
        (DE[i, j, t] == d_estimada[f"{i},{j},{t}"] + gamma[i, j] * (p[i, j, t] - p[i, j, t - 1])
        for i in I_ for j in J_ for t in T_ if t > prm.T_INICIO), name="def_DE_rest"
    )



    # 6. Qijt >= CMOi	* z_ijt	
    m.addConstrs(q[i, j, t] >= CMO[i] * z[i, j, t] for i in I_ for j in J_ for t in T_)

    # 7. Demanda insatisfecha  DIijt(Pijt) = Dijt(Pijt) – Vijt 
    m.addConstrs(di[i, j, t] == d[i, j, t] - v[i, j, t] for i in I_ for j in J_ for t in T_)

    # 8. Venta del producto Vijt <= Dijt(pijt), Vijt <= Iijt
    m.addConstrs(v[i, j, t] <= d[i, j, t] for i in I_ for j in J_ for t in T_)
    m.addConstrs(v[i, j, t] <= inv[i, j, t] for i in I_ for j in J_ for t in T_)
  

    # 9. Iijt+1 = Iijt - Vijt + Qijt	
    m.addConstrs(inv[i, j, t] == inv[i, j, t-1] - v[i, j, t] + q[i, j, t-1] for i in I_ for j in J_ for t in T_[1:])
    m.addConstrs(inv[i, j, prm.T_INICIO] == 5000 for i in I_ for j in J_)   

# DEBERÍA PODER AGREGARSE LA RESTRICCIÓN YA QUE YA SE TIENE A d, pero no tenemos A ni B!
    # 10 pijt = Aijt + B1ijt * DEijt + B2ijt * DEijt^2 
   # m.addConstrs(
    #p[i, j, t] <= max_(
     #   A[i, j] + B1[i, j] * DE[i, j, t] + B2[i, j] * DE[i, j, t] * DE[i, j, t],
      #  proporcion_maxima * C_PROD[i]
    #)
    #for i in I_ for j in J_ for t in T_
    #)

    ## no deja usar max asi que modifico
    # crea la variable auxiliar (esto va en tu clase de variables)
    p_max = m.addVars(I_, J_, T_, name="p_max")

    # define que sea el máximo
    m.addConstrs(
        (p_max[i, j, t] >= A[i, j] + B1[i, j] * DE[i, j, t] + B2[i, j] * DE[i, j, t] * DE[i, j, t]
        for i in I_ for j in J_ for t in T_),
        name="p_max_expr"
    )
    m.addConstrs(
        (p_max[i, j, t] >= proporcion_maxima * C_PROD[i]
        for i in I_ for j in J_ for t in T_),
        name="p_max_cap"
    )

    # finalmente, usa esa variable en la restricción original
    m.addConstrs(
        (p[i, j, t] <= p_max[i, j, t]
        for i in I_ for j in J_ for t in T_),
        name="p_limited_by_max"
    )

    m.update()

    m.setParam('MIPGap', 0.01) ## No achicar, con gap 0.001 tarda > 6 horas
    m.setParam('NonConvex', 2)
    print("Tamaño del modelo:")
    print(f"  Productos: {len(modelo.I_)}")
    print(f"  Tiendas: {len(modelo.J_)}")
    print(f"  Semanas: {len(modelo.T_)}")
    print(f"  Variables: {modelo.m.NumVars}")
    print(f"  Restricciones: {modelo.m.NumConstrs}")
    m.optimize()
    # print(f"Valor óptimo de la función objetivo: {m.ObjVal}")
    df_resultados = exportar_resultados_a_df(modelo)
    exportar_a_bbdd(modelo)

    #if m.status == GRB.OPTIMAL:
    #    print(f"Valor óptimo de la función objetivo: {m.ObjVal:.2f}")
    #    for i in I_:
    #        for j in J_:
    #            for t in T_:
    #                print(f"Producto {i}, Tienda {j}, Semana {t}:")
    #                print(f"  Precio: {p[i,j,t].X:.2f}")
    #                print(f"  Pedido: {q[i,j,t].X:.0f}")
    #                print(f"  Venta: {v[i,j,t].X:.0f}")
    #                print(f"  Inventario: {inv[i,j,t].X:.0f}")
    #                print(f"  Demanda insatisfecha: {di[i,j,t].X:.0f}")
    #else:
    #    print("El modelo no encontró solución óptima.")

    insatisfecha_total = df_resultados["DemandaInsatisfecha"].sum()
    demanda_total = df_resultados["DemandaInsatisfecha"].sum() + df_resultados["Venta"].sum()
    proporcion_insatisfecha = float(insatisfecha_total / demanda_total)

    return [m.ObjVal, demanda_total, insatisfecha_total, proporcion_insatisfecha]

def exportar_resultados_a_df(modelo: ModeloVariablesParametros):
    q, p, z, v, inv, di, DE = modelo.obtener_variables()
    I_, J_, T_ = modelo.obtener_rangos().values()

    resultados = []
    for i in I_:
        for j in J_:
            for t in T_:
                resultados.append({
                    'Producto': i,
                    'Tienda': j,
                    'Semana': t,
                    'Precio': p[i, j, t].X,
                    'Pedido': q[i, j, t].X,
                    'Venta': v[i, j, t].X,
                    'Inventario': inv[i, j, t].X,
                    'DemandaInsatisfecha': di[i, j, t].X,
                    'Activacion': z[i, j, t].X
                })

    df = pd.DataFrame(resultados)
    df.to_csv("resultados_modelo.csv", index=False)
    # print("Resultados guardados en resultados_modelo.csv")
    return df

def exportar_a_bbdd(modelo: ModeloVariablesParametros):
    q, p, z, v, inv, di, DE = modelo.obtener_variables()
    I_, J_, T_ = modelo.obtener_rangos().values()
    
    # Crear listas con los resultados de ambas tiendas, por producto, en formato de los datos
    resultados_t1 = []
    resultados_t2 = []
    for t in T_:
        resultados_t1.append({
            'Semana': t,
            'D1': v[1, 1, t].X + di[1, 1, t].X,
            'P1': p[1, 1, t].X,
            'D2': v[2, 1, t].X + di[2, 1, t].X,
            'P2': p[2, 1, t].X,
            'D3': v[3, 1, t].X + di[3, 1, t].X,
            'P3': p[3, 1, t].X,
            'D4': v[4, 1, t].X + di[4, 1, t].X,
            'P4': p[4, 1, t].X,
            'D5': v[5, 1, t].X + di[5, 1, t].X,
            'P5': p[5, 1, t].X,
            'D6': v[6, 1, t].X + di[6, 1, t].X,
            'P6': p[6, 1, t].X,
            'D7': v[7, 1, t].X + di[7, 1, t].X,
            'P7': p[7, 1, t].X,
            'D8': v[8, 1, t].X + di[8, 1, t].X,
            'P8': p[8, 1, t].X,
            'D9': v[9, 1, t].X + di[9, 1, t].X,
            'P9': p[9, 1, t].X,
            'D10': v[10, 1, t].X + di[10, 1, t].X,
            'P10': p[10, 1, t].X,
        })
        resultados_t2.append({
            'Semana': t,
            'D1': v[1, 2, t].X + di[1, 2, t].X,
            'P1': p[1, 2, t].X,
            'D2': v[2, 2, t].X + di[2, 2, t].X,
            'P2': p[2, 2, t].X,
            'D3': v[3, 2, t].X + di[3, 2, t].X,
            'P3': p[3, 2, t].X,
            'D4': v[4, 2, t].X + di[4, 2, t].X,
            'P4': p[4, 2, t].X,
            'D5': v[5, 2, t].X + di[5, 2, t].X,
            'P5': p[5, 2, t].X,
            'D6': v[6, 2, t].X + di[6, 2, t].X,
            'P6': p[6, 2, t].X,
            'D7': v[7, 2, t].X + di[7, 2, t].X,
            'P7': p[7, 2, t].X,
            'D8': v[8, 2, t].X + di[8, 2, t].X,
            'P8': p[8, 2, t].X,
            'D9': v[9, 2, t].X + di[9, 2, t].X,
            'P9': p[9, 2, t].X,
            'D10': v[10, 2, t].X + di[10, 2, t].X,
            'P10': p[10, 2, t].X,
        })

    # Crear los df de ambos resultados
    df_t1 = pd.DataFrame(resultados_t1)
    df_t2 = pd.DataFrame(resultados_t2)

    # Crear archivos Excel
    file_t1 = 'resultados_t1.xlsx'
    file_t2 = 'resultados_t2.xlsx'

    # Comprobar existencia de los archivos Excel y escrubir df
    if not os.path.exists(file_t1):
        with pd.ExcelWriter(file_t1, engine = 'openpyxl') as writer:
            df_t1.to_excel(writer, index = False, header = True, sheet_name = 'Hoja1')
    else:
        with pd.ExcelWriter(file_t1, mode = 'a', engine = 'openpyxl',
                            if_sheet_exists = 'overlay') as writer:
            hoja_t1 = writer.sheets['Hoja1']
            startrow_t1 = hoja_t1.max_row
            df_t1.to_excel(writer, index = False, header = False, sheet_name = 'Hoja1',
                           startrow = startrow_t1)
    
    if not os.path.exists(file_t2):
        with pd.ExcelWriter(file_t2, engine = 'openpyxl') as writer:
            df_t2.to_excel(writer, index = False, header = True, sheet_name = 'Hoja1')
    else:
        with pd.ExcelWriter(file_t2, mode = 'a', engine = 'openpyxl',
                            if_sheet_exists = 'overlay') as writer:
            hoja_t2 = writer.sheets['Hoja1']
            startrow_t2 = hoja_t2.max_row
            df_t2.to_excel(writer, index = False, header = False, sheet_name = 'Hoja1',
                           startrow = startrow_t2)
            
def exportar_a_bbdd(modelo):
    q, p, z, v, inv, di, DE = modelo.obtener_variables()
    I_, J_, T_ = modelo.obtener_rangos().values()

    resultados_t1, resultados_t2 = [], []
    for t in T_:
        fila_t1, fila_t2 = {"Semana": t}, {"Semana": t}
        for i in range(1, 11):
            fila_t1[f"D{i}"] = v[i, 1, t].X + di[i, 1, t].X
            fila_t1[f"P{i}"] = p[i, 1, t].X
            fila_t2[f"D{i}"] = v[i, 2, t].X + di[i, 2, t].X
            fila_t2[f"P{i}"] = p[i, 2, t].X
        resultados_t1.append(fila_t1)
        resultados_t2.append(fila_t2)

    df_t1 = pd.DataFrame(resultados_t1)
    df_t2 = pd.DataFrame(resultados_t2)

    guardar_con_seguridad("resultados_t1.xlsx", df_t1)
    guardar_con_seguridad("resultados_t2.xlsx", df_t2)

def guardar_con_seguridad(nombre, df):
        try:
            if os.path.exists(nombre):
                try:
                    load_workbook(nombre)
                except Exception:
                    os.remove(nombre)
            with pd.ExcelWriter(nombre, engine='openpyxl', mode='a' if os.path.exists(nombre) else 'w') as writer:
                df.to_excel(writer, index=False, sheet_name='Hoja1')
        except Exception as e:
            print(f"Error al guardar {nombre}: {e}")


if __name__ == "__main__":
    resultado = correr_modelo()
    print("Resultado:", resultado)